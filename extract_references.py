#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Extract ASHRAE 140 BESTEST reference data from RESULTS5-2A-Update.xlsx.

Reads the Excel file and produces consolidated TSV reference files:
  - data/reference/annual-references.tsv       (annual heating/cooling + peak loads)
  - data/reference/free-float-references.tsv    (free-float max/min/avg temperatures)
  - data/reference/monthly-references.tsv       (monthly heating & cooling min/max, Case 600 & 900)

Run once:
    python extract_references.py
"""

from __future__ import annotations

import re
from pathlib import Path

import openpyxl
import pandas as pd


EXCEL_PATH = Path("validation_results/RESULTS5-2A-Update.xlsx")
OUT_DIR = Path("data/reference")


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _extract_case_number(label: str) -> str | None:
    """Extract the case number/label from an Excel row label string.

    Examples:
        '600 Base Case, South Windows' -> '600'
        '600FF - Low Mass with S. Windows' -> '600FF'
        '650FF Case 600FF with Night Ventilation' -> '650FF'
        '685 Case 600 with "20/20" Thermostat ' -> '685'
        '960 Sunspace ' -> '960'
    """
    if not label or not isinstance(label, str):
        return None
    label = label.strip()
    m = re.match(r"(\d+(?:FF)?)\b", label)
    return m.group(1) if m else None


def _read_block(ws, first_data_row: int, last_data_row: int,
                case_col: int, min_col: int, max_col: int) -> dict[str, tuple[float | None, float | None]]:
    """Read a block of rows from a worksheet and return {case: (min, max)}."""
    result: dict[str, tuple[float | None, float | None]] = {}
    for r in range(first_data_row, last_data_row + 1):
        label = ws.cell(row=r, column=case_col).value
        case = _extract_case_number(label)
        if not case:
            continue
        v_min = ws.cell(row=r, column=min_col).value
        v_max = ws.cell(row=r, column=max_col).value
        # Treat None / -1 as missing
        if v_min is not None and v_min != -1:
            v_min = float(v_min)
        else:
            v_min = None
        if v_max is not None and v_max != -1:
            v_max = float(v_max)
        else:
            v_max = None
        result[case] = (v_min, v_max)
    return result


def _read_monthly_block(ws, header_row: int, case_col: int, min_col: int, max_col: int,
                        cases: list[tuple[str, int]]) -> dict[str, list[tuple[float | None, float | None]]]:
    """Read monthly blocks (12 rows per case) and return {case: [(min,max) x 12]}."""
    result = {}
    for case_label, start_row in cases:
        months = []
        for offset in range(12):
            r = start_row + offset
            v_min = ws.cell(row=r, column=min_col).value
            v_max = ws.cell(row=r, column=max_col).value
            if v_min is not None:
                v_min = float(v_min)
            if v_max is not None:
                v_max = float(v_max)
            months.append((v_min, v_max))
        result[case_label] = months
    return result


# ---------------------------------------------------------------------------
# Main extraction
# ---------------------------------------------------------------------------

def main() -> None:
    print(f"Reading {EXCEL_PATH} ...")
    wb = openpyxl.load_workbook(str(EXCEL_PATH), data_only=True)

    # -----------------------------------------------------------------------
    # 1. Annual heating & cooling  (Tables 1, Tables B8-1 and B8-2)
    # -----------------------------------------------------------------------
    ws1 = wb["Tables 1"]

    # Table B8-1: Annual Heating (rows 11-56), cols: B(2)=label, K(11)=Min, L(12)=Max
    heating = _read_block(ws1, 11, 56, case_col=2, min_col=11, max_col=12)

    # Table B8-2: Annual Cooling (rows 63-108), same col layout
    cooling = _read_block(ws1, 63, 108, case_col=2, min_col=11, max_col=12)

    # -----------------------------------------------------------------------
    # 2. Peak heating & cooling  (Tables 2, Tables B8-3 and B8-4)
    # -----------------------------------------------------------------------
    ws2 = wb["Tables 2"]

    # Table B8-3: Peak Heating (rows 11-56), cols: B(2)=label, AI(35)=Min, AJ(36)=Max
    peak_heating = _read_block(ws2, 11, 56, case_col=2, min_col=35, max_col=36)

    # Table B8-4: Peak Cooling (rows 63-108)
    peak_cooling = _read_block(ws2, 63, 108, case_col=2, min_col=35, max_col=36)

    # -----------------------------------------------------------------------
    # 3. Free-float temperatures  (Tables 2, Table B8-5)
    # -----------------------------------------------------------------------
    # Max temperature: rows 116-122
    ff_max = _read_block(ws2, 116, 122, case_col=2, min_col=35, max_col=36)
    # Min temperature: rows 127-133
    ff_min = _read_block(ws2, 127, 133, case_col=2, min_col=35, max_col=36)
    # Average temperature: rows 138-144
    ff_avg = _read_block(ws2, 138, 144, case_col=2, min_col=35, max_col=36)

    # -----------------------------------------------------------------------
    # 4. Monthly heating & cooling  (Tables M1, Tables B8-M1 and B8-M2)
    # -----------------------------------------------------------------------
    wsm = wb["Tables M1"]

    # Monthly Heating: Case 600 starts at row 11, Case 900 starts at row 24
    monthly_heating = _read_monthly_block(
        wsm, header_row=7, case_col=2, min_col=11, max_col=12,
        cases=[("600", 11), ("900", 24)],
    )

    # Monthly Cooling: Case 600 starts at row 42, Case 900 starts at row 55
    monthly_cooling = _read_monthly_block(
        wsm, header_row=38, case_col=2, min_col=11, max_col=12,
        cases=[("600", 42), ("900", 55)],
    )

    # -----------------------------------------------------------------------
    # Write annual-references.tsv
    # -----------------------------------------------------------------------
    all_cases = sorted(
        set(heating.keys()) | set(cooling.keys()),
        key=lambda c: (int(re.sub(r"FF$", "", c)), "FF" in c),
    )
    rows_annual = []
    for case in all_cases:
        h_min, h_max = heating.get(case, (None, None))
        c_min, c_max = cooling.get(case, (None, None))
        ph_min, ph_max = peak_heating.get(case, (None, None))
        pc_min, pc_max = peak_cooling.get(case, (None, None))
        rows_annual.append({
            "Case": case,
            "heating_min": h_min,
            "heating_max": h_max,
            "cooling_min": c_min,
            "cooling_max": c_max,
            "peak_heating_min": ph_min,
            "peak_heating_max": ph_max,
            "peak_cooling_min": pc_min,
            "peak_cooling_max": pc_max,
        })
    df_annual = pd.DataFrame(rows_annual)
    out_annual = OUT_DIR / "annual-references.tsv"
    df_annual.to_csv(out_annual, sep="\t", index=False, float_format="%.6f")
    print(f"  Written {out_annual}  ({len(df_annual)} cases)")

    # -----------------------------------------------------------------------
    # Write free-float-references.tsv
    # -----------------------------------------------------------------------
    ff_cases = sorted(
        set(ff_max.keys()) | set(ff_min.keys()) | set(ff_avg.keys()),
        key=lambda c: (int(re.sub(r"FF$", "", c)), "FF" in c),
    )
    rows_ff = []
    for case in ff_cases:
        tmax_min, tmax_max = ff_max.get(case, (None, None))
        tmin_min, tmin_max = ff_min.get(case, (None, None))
        tavg_min, tavg_max = ff_avg.get(case, (None, None))
        rows_ff.append({
            "Case": case,
            "temp_max_min": tmax_min,
            "temp_max_max": tmax_max,
            "temp_min_min": tmin_min,
            "temp_min_max": tmin_max,
            "temp_avg_min": tavg_min,
            "temp_avg_max": tavg_max,
        })
    df_ff = pd.DataFrame(rows_ff)
    out_ff = OUT_DIR / "free-float-references.tsv"
    df_ff.to_csv(out_ff, sep="\t", index=False, float_format="%.6f")
    print(f"  Written {out_ff}  ({len(df_ff)} cases)")

    # -----------------------------------------------------------------------
    # Write monthly-references.tsv  (replaces existing file)
    # -----------------------------------------------------------------------
    month_labels = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
                    "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
    monthly_cases = sorted(
        set(monthly_heating.keys()) | set(monthly_cooling.keys()),
        key=lambda c: int(c),
    )

    # Build column list
    cols = ["Month"]
    for case in monthly_cases:
        cols.extend([
            f"Case{case}_heating_min", f"Case{case}_heating_max",
            f"Case{case}_cooling_min", f"Case{case}_cooling_max",
        ])

    rows_m: list[dict] = []
    for i, month in enumerate(month_labels):
        row: dict = {"Month": month}
        for case in monthly_cases:
            h = monthly_heating.get(case, [(None, None)] * 12)
            c = monthly_cooling.get(case, [(None, None)] * 12)
            h_min, h_max = h[i]
            c_min, c_max = c[i]
            row[f"Case{case}_heating_min"] = round(h_min, 1) if h_min is not None else ""
            row[f"Case{case}_heating_max"] = round(h_max, 1) if h_max is not None else ""
            row[f"Case{case}_cooling_min"] = round(c_min, 1) if c_min is not None else ""
            row[f"Case{case}_cooling_max"] = round(c_max, 1) if c_max is not None else ""
        rows_m.append(row)

    df_m = pd.DataFrame(rows_m, columns=cols)
    out_m = OUT_DIR / "monthly-references.tsv"
    df_m.to_csv(out_m, sep="\t", index=False)
    print(f"  Written {out_m}  ({len(monthly_cases)} cases x 12 months)")

    print("Done.")


if __name__ == "__main__":
    main()
